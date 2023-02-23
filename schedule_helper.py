# ====IMPORTS====
import os
import re
import sys
import csv
import eel
import time
import shutil
import platform
import openpyxl 
import subprocess 

from bs4 import BeautifulSoup

from datetime import timedelta
from openpyxl.styles import Border, Side

import ebooklib
from ebooklib import epub

# ===============

# Finds all files with the .epub extension in the /epubs directory
epubs = [e for e in os.listdir('epubs') if e.endswith('.epub')]
if(len(epubs) != 1): # Raises an error if more than one .epub file is present
    if(len(epubs) < 1):
        print('No epub files present')
    else:
        print('Too many epub files in the folder.\nRemove all but the one you wish to convert')
    time.sleep(5) # Sleep to allow time for the user to see the error
    raise ValueError("Incorrect number of files")


bookPath = epubs[0]
book = epub.read_epub(f'epubs/{bookPath}')
# shutil.move(f'epubs/{bookPath}', f'epubs archive/{bookPath}')


# Fills the bookText list with the raw item content from the .epub
bookText = []
for item in book.get_items():
    if item.get_type() == ebooklib.ITEM_DOCUMENT:
        bookText.append(item.get_content())

# Opens a textfile with utf-8 encoding to deal with problematic special characters
file = open('src/sections.txt', 'w', encoding="utf-8")

dates = [] # Used to store the date ranges for each individual week
scriptures = [] # Used to store the scriptures included at the top of each week

months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

# Uses bs4 to loop through each chunk of binary data from the epub end interpret it
for chunk in bookText:
    soup = BeautifulSoup(chunk, 'html.parser')
    
    # Extracts the relavent body text from the epub, stored within section divs
    for section in soup.find_all('div', {'class' : 'section'}):
        file.write(section.get_text())
    
    # Extracts text containing the scripture associated with each week's sheet
    for section in soup.find_all('strong'):
        for subSection in section.find_all('a', href="#citation1"):
            result = re.search("\d*\s*[A-Z]+\s\d+-\d+|\d*\s*[A-Z]+\s\d+", subSection.get_text())
            if result != None:
                # Removes unreadable character so that the result can be stored as plain text
                result = re.sub("\xa0", " ", result.group()) 
                scriptures.append(result)

    # Extracts text containing the date range associated with each week's sheet
    for p in soup.find_all('span', {'class' : 'pageNum'}):
        for month in months: 
            value = p.next_sibling.get_text()

            if month in value and "Workbook" not in value:
                test = re.search(".*\s\d+[-–]+.*\d+", value)
                
                if(test != None):
                    result = re.sub("\xa0", " ", test.group())
                    result = re.sub("-|–", " - ", result)
                    if(result not in dates):
                        dates.append(result)
file.close()


file = open('src/sections.txt', 'r', encoding="utf-8")
sectionText = file.read()
file.close()

# Regular expressions that remove problematic characters and repalces
# a newline character that interferes with value extraction
x = re.sub("\u200b", " ", sectionText)
y = re.sub("\xa0", " ", x)
z = re.sub("Spiritual\sGems:\s\(10\smin.\)\n", "Spiritual Gems: (10 min.) ", y)

# Opens the file again without utf-8 encoding, to ensure that it can be opened
# and read in the future as plain text without issue
file = open('src/sections.txt', 'w+')
file.write(z)
file.close()

file = open('src/sections.txt', 'r')
text = file.read()

# Finds every song referenced in the text and appends it to the songs list
songs = [] 
result = re.findall("Song \d+", text)
for song in result:
    songs.append(song)

talks = [] 
talkTime = {} # Dictionary containing pairs of talks and their associated length in minutes 
#Finds each talk and time indicator that needs to be extracted  
id = 0
result = re.findall(".*\(\d+\smin.\).*\n|Song \d+", text)
for talk in result: 
    result = re.search("\(\d+\smin.\)", talk)
    removeRe = re.sub("\(\d+\smin.\)", "", talk)

    if("Opening Comments" in removeRe or "Concluding Comments" in removeRe):
        removeRe += str(id)
        id += 1

    talks.append(removeRe)

    if result is not None:
        dotResult = re.sub("\.", "", result.group())
        talkTime[removeRe] = dotResult
file.close()

# Populating lists with the talks associated with each section of the sheet 
ministry = [[]]
treasures = [[]]
christians = [[]]

mIndex = 0
tIndex = 0
cIndex = 0

current = 'treasures'

for i in range(len(talks)):
    if current == 'christians' and 'Song' not in talks[i]:
        christians[cIndex].append(talks[i])
        if("Song" in talks[i + 1]):
            christians.append([])
            cIndex += 1
            current = 'treasures'
    if current == 'ministry' and 'Song' not in talks[i] and 'Concluding Comments' not in talks[i]:
        ministry[mIndex].append(talks[i])
        if("Song" in talks[i + 1]):
            ministry.append([])
            mIndex += 1
            current = 'christians'
    if current == 'treasures' and 'Song' not in talks[i] and 'Concluding Comments' not in talks[i]:
        treasures[tIndex].append(talks[i])
        if "Bible Reading:" in talks[i]:
            treasures.append([])
            tIndex += 1
            current = 'ministry'

ministry.pop()
treasures.pop()
christians.pop()



# ====USER-INTERFACE====
# root = Tk()
# root.geometry('800x500')
# root.title('Speaker assignment')

# # Create a main frame
# main_frame = Frame(root)
# main_frame.pack(fill=BOTH, expand=1)

# # Create a canvas, so that the entire window can be scrolled
# my_canvas = Canvas(main_frame)
# my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

# # Add a scrollbar to the Canvas
# my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
# my_scrollbar.pack(side=RIGHT, fill=Y)

# # configure the canvas
# my_canvas.configure(yscrollcommand=my_scrollbar.set)
# my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox('all')))

# # create another frame inside the canvas
# second_frame = Frame(my_canvas)
# my_canvas.create_window((0,0), window=second_frame, anchor="nw")


# dateIndex = 0
talkSpeakerDict = {} # Dictionary that associates a speaker value with a talk key
headingColor = ''

eel.init(f'{os.path.dirname(os.path.realpath(__file__))}/web')

# displayTalks = []

# for i in range(len(talks)):
#     if "Song" not in talks[i]:
#         # Truncates string if it contains 85 characters or greater
#         try:
#             result = re.search(".{85}", talks[i]).group()
#             result += "..." 
#         except:
#             result = talks[i]

#         displayTalks.append(result)


names = []
with open('src/names.csv') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    for row in csv_reader:
        names.append(row[0])
names = names[1:]
print(names)

eel.setup(talks, dates, names)


open = True
@eel.expose
def end_program():
    global open
    open = False
    print("test")

@eel.expose
def take_input(values, color=''):
    global headingColor
    print(values)
    for value in values:
        talkSpeakerDict[value[1]] = value[0]
    headingColor = color[1:]

    print(headingColor.upper())
    # headingColor = headingColor[1:].upper()
    # headingColor = re.sub('\D', '', headingColor)
    

    # print(rgb_to_hex_conversion(headingColor[0:1], headingColor[2:3], headingColor[4:5]))

@eel.expose
def open_directory():
    subprocess.Popen(f'explorer "{os.path.abspath("./epubs")}"')

@eel.expose
def add_speaker(value):
    print(value)
    file = os.open("src/names.csv", os.O_APPEND)
    writer = csv.writer(file)
    writer.writerow([[value],[]]) 
    # ???????????????

        # with open('src/names.csv', 'w') as csv_file:
    #     writer = csv.writer(csv_file)
    #     writer.writerow([f'{value}'])


eel.start("index.html", size=(1100, 800), block=False)

while True:
    
    eel.sleep(1.0)

    if(open == False):
        break
    #do things


print("out of loops")
print(headingColor)


# # Captures the value of the Entry widgets when this function is called
# def captureInput(event, talk):
#     value = event.widget.get()
#     talkSpeakerDict[talk] = value


# for i in range(len(talks)):
#     # Intermediary handler for each Entry widget created so that unique values can be passed
#     def handler(event, talkIndex=i):
#             captureInput(event, talks[talkIndex])


#     if "Song" not in talks[i]:
#         # Truncates string if it contains 85 characters or greater
#         try:
#             result = re.search(".{85}", talks[i]).group()
#             result += "..." 
#         except:
#             result = talks[i]

#         # Creates a label and associated Entry field for the user to input 
#         # a speaker to be assigned to a specific talk
#         if("Opening Comments" not in result):
#             Label(second_frame, text=result, anchor="e", width=75).grid(row=i, column=0, pady=10, padx=0)

#             entry = (Entry(second_frame))
#             entry.grid(row=i, column=1, pady=10, padx=10)
#             entry.bind('<KeyRelease>', handler)
#         else: # Creates a label holding the date range of the sheet in place of the sheet
#             Label(second_frame, text=dates[dateIndex], anchor="e", width=75).grid(row=i, column=0, pady=10, padx=0)
#             dateIndex += 1
dateIndex = 0

# def close():
#     root.quit()
# Button(second_frame, text='Apply', command=close).grid(row=1, column=3)

# root.mainloop()


# ====SPREADSHEET-EDITING====
wb = openpyxl.load_workbook('src/template.xlsx')
ws1 = wb.active
rows = ws1.max_row

# Replaces each instance of [Song] in the template sheet
# with the next item in the songs list
songCounter = 0
for row in range(1, rows + 1):
    if(ws1[f'A{row}'].value == '[Song]'):
        ws1[f'A{row}'] = songs[songCounter]
        
        if(songCounter < len(songs) - 1):
            songCounter += 1
        

mIndex = 0
tIndex = 0
cIndex = 0
current = 'treasures'

def insertTalk(row):
    global rows

    rows += 1
    ws1.insert_rows(row, 1)

    ws1[f'C{row}'] = talk 
    ws1[f'A{row}'] = talkTime[talk]   

    ws1[f'A{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    if(len(ws1[f'C{row}'].value) > 65):
        ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=True)
    else:
        ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=False)

    if(talk in talkSpeakerDict):
        ws1[f'B{row}'] = talkSpeakerDict[talk]
        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1[f'B{row}'].font = openpyxl.styles.Font(bold=True)


row = 1
while row < rows:
# for row in range(1, 1000):
    if(ws1[f'C{row}'].value == '[Talk]'):
        if(current == 'treasures'):
            ws1[f'C{row}'] = ""
            for talk in reversed(treasures[tIndex]):
                insertTalk(row)
            
            if(tIndex < len(treasures) - 1):
                tIndex += 1
            current = 'ministry'
            continue

        if(current == 'ministry'):
            ws1[f'C{row}'] = ""
            for talk in reversed(ministry[mIndex]):
                insertTalk(row)

            if(mIndex < len(ministry) - 1):
                mIndex += 1
            current = 'christians'
            continue

        if(current == 'christians'):
            ws1[f'C{row}'] = ""
            for talk in reversed(christians[cIndex]):
                insertTalk(row)
            
            if(cIndex < len(christians) - 1):
                cIndex += 1
            current = 'treasures'
    row += 1


top = Side(border_style='thin', color='FFFFFF')
left = Side(border_style='thin', color='FFFFFF')
right = Side(border_style='thin', color='FFFFFF')
bottom = Side(border_style='thin', color='FFFFFF')
border = Border(top = top, bottom = bottom, left = left, right = right)

cellRange = ws1['A1' : f'D{rows}']
for cell in cellRange:
    for x in cell:
        x.border = border

bottom = Side(border_style='thin', color='808080')
border = Border(top = top, bottom = bottom, left = left, right = right)


finishTime = timedelta(0, (7*3600+6*60))


row = 1
while row < rows:
    if(dateIndex >= len(dates)):  
        break

    if(ws1[f'B{row}'].value == '[Date]'):
        ws1[f'B{row}'] = dates[dateIndex]
        dateIndex+=1

        if(len(ws1[f'B{row}'].value) > 16):
            ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center')
        else:
            ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(wrap_text=False, horizontal='center')

    value = ws1[f'C{row}'].value
    if(value != None and "Closing Prayer" in value): 
        ws1.delete_rows(row - 1, 1)
        rows -= 1


    value = ws1[f'A{row}'].value

    if(ws1[f'C{row + 1}'].value != None and 'Opening Prayer' in ws1[f'C{row + 1}'].value):
        headerFill = openpyxl.styles.PatternFill(patternType='solid', fill_type='solid', fgColor=openpyxl.styles.Color(headingColor))
        ws1[f'A{row}'].fill = headerFill
        ws1[f'B{row}'].fill = headerFill
        ws1[f'C{row}'].fill = headerFill
        ws1[f'D{row}'].fill = headerFill
    
    if(value != None and 'min' in value):
        stripNum = re.search('\d+', ws1[f'A{row}'].value)
        stripNum = int(stripNum.group())


        cellRange = ws1[f'A{row}' : f'D{row}']

        if('th study' in ws1[f'C{row}'].value):
            stripNum += 1
            if('Bible Reading' not in ws1[f'C{row}'].value):
                ws1.insert_rows(row + 1, 1)
                rows += 1
                
                if(ws1[f'C{row}'].value + '-householder' in talkSpeakerDict):
                    ws1[f'B{row + 1}'] = talkSpeakerDict[ws1[f'C{row}'].value + '-householder']
                    ws1[f'B{row + 1}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    ws1[f'B{row + 1}'].font = openpyxl.styles.Font(bold=True)
                
                grayFill = openpyxl.styles.PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
                ws1[f'B{row + 1}'].fill = grayFill
                cellRange = ws1[f'A{row + 1}' : f'D{row + 1}']

        
        for cell in cellRange:
            for x in cell:
                x.border = border

        timeAdd = timedelta(0, (0*3600+stripNum)*60)

        finishTime += timeAdd 
        
        if(ws1[f'C{row + 1}'].value != None and "Congregation Bible Study: " in ws1[f'C{row + 1}'].value):
            finishTime = timedelta(0, (8*3600+7*60))

        ws1[f'D{row}'] = finishTime
        
        number_format = '[hh]:mm'
        ws1[f'D{row}'].number_format = number_format
        ws1[f'D{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        if("Concluding Comments" in ws1[f'C{row}'].value):
            finishTime = timedelta(0, (7*3600+6*60))

    row += 1




first = re.sub("-.*\s", " ", book.get_metadata('DC', 'title')[0][0])
second = re.sub(",\s.*-", ", ", book.get_metadata('DC', 'title')[0][0])
ws1['B1'] = first


secondCoord = None
scripturesIndex = 0

for row in range(2, rows):
    if(ws1[f'C{row}'].value != None):
        if("Opening Comments" in ws1[f'C{row}'].value or "Concluding Comments" in ws1[f'C{row}'].value):
            stripNum = re.sub('\d+', " ", ws1[f'C{row}'].value)
            ws1[f'C{row}'] = stripNum

    if(ws1[f'B{row}'].value == '[Heading]'):
        secondCoord = row
        ws1[f'B{row}'] = second

        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1.merge_cells(f'B{row}:C{row}')

    if(ws1[f'C{row}'].value == '[Scripture]' and scripturesIndex < len(scriptures) ):
        ws1[f'C{row}'] = "   " + scriptures[scripturesIndex]
        ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(vertical='center')
        scripturesIndex += 1


cellRange = ws1[f'D2' : f'D{secondCoord - 3}']

for i in range(2):
    for cell in cellRange:
        for x in cell:
            top = Side(**x.border.top.__dict__)
            left = Side(**x.border.left.__dict__)
            bottom = Side(**x.border.bottom.__dict__)

            right = Side(border_style='thin', color='808080')
            border = Border(top = top, bottom = bottom, left = left, right = right)
            x.border = border

    cellRange = ws1[f'D{secondCoord + 1}' : f'D{rows}']
cellRange = ws1[f'A{secondCoord - 3}' : f'D{secondCoord - 3}']



for i in range(2):
    for cell in cellRange:
        for x in cell:
            top = Side(**x.border.top.__dict__)
            left = Side(**x.border.left.__dict__)
            right = Side(**x.border.right.__dict__)

            bottom = Side(border_style='thin', color='808080')
            border = Border(top = top, bottom = bottom, left = left, right = right)
            x.border = border

    cellRange = ws1[f'A{rows}' : f'D{rows}']


wb.save(filename = f"output/{book.title}.xlsx")


# os.chdir('./output')

# if(platform.system() == "Windows"):
#     os.startfile(f'{book.title}.xlsx')
# else:
#     opener ="open" if sys.platform == "darwin" else "xdg-open"
#     subprocess.call([opener, f'{book.title}.xlsx'])



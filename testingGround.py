import ebooklib
from ebooklib import epub
from bs4 import BeautifulSoup
import os
import re 
from pprint import pprint







os.system('cls')
book = epub.read_epub('ref.epub')


fileText = []
scriptures = []

for item in book.get_items():
    if item.get_type() == ebooklib.ITEM_DOCUMENT:
        fileText.append(item.get_content())
        

file = open('sections.txt', 'w', encoding="utf-8")

for snippet in fileText:
    soup = BeautifulSoup(snippet, 'html.parser')
    
    for section in soup.find_all('div', {'class' : 'section'}):
        file.write(section.get_text())

    
    for test in soup.find_all('strong'):
        for section in test.find_all('a', href="#citation1"):
            test = re.sub("\xa0", " ", section.get_text())
            scriptures.append(test)


 
pprint(scriptures)

file.close()


file = open('sections.txt', 'r', encoding="utf-8")

sectionText = file.read()
x = re.sub("\u200b", " ", sectionText)
y = re.sub("\xa0", " ", x)
z = re.sub("Spiritual\sGems:\s\(10\smin.\)\n", "Spiritual Gems: (10 min.) ", y)

file.close()


file = open('sections.txt', 'w+')
# file.write(y)
file.write(z)
file.close()


file = open('sections.txt', 'r')
text = file.read()
x = re.findall("Song \d+", text)

songs = [] 
for song in x:
    songs.append(song)

pprint(songs)

talks = []
talkTime = {}


x = re.findall(".*\(\d+\smin.\).*\n|Song \d+", text)
for talk in x: 
    testRe = re.search("\(\d+\smin.\)", talk)

    removeRe = re.sub("\(\d+\smin.\)", "", talk)
    talks.append(removeRe)

    if testRe is not None:
        dotRemRe = re.sub("\.", "", testRe.group())
        talkTime[removeRe] = dotRemRe
    
    
file.close()

pprint(talkTime)


ministry = [[]]
treasures = [[]]
christians = [[]]

mIndex = 0
tIndex = 0
cIndex = 0

current = 'treasures'

for i in range(len(talks)):
    if current == 'christians' and 'Song' not in talks[i]:
        christians[cIndex].append(talks[i])
        if("Song" in talks[i + 1]):
            christians.append([])
            cIndex += 1
            current = 'treasures'
    if current == 'ministry' and 'Song' not in talks[i] and 'Concluding Comments' not in talks[i]:
        ministry[mIndex].append(talks[i])
        if("Song" in talks[i + 1]):
            ministry.append([])
            mIndex += 1
            current = 'christians'
    if current == 'treasures' and 'Song' not in talks[i] and 'Concluding Comments' not in talks[i]:
        treasures[tIndex].append(talks[i])
        if "Bible Reading:" in talks[i]:
            treasures.append([])
            tIndex += 1
            current = 'ministry'

ministry.pop()
treasures.pop()
christians.pop()

pprint(treasures)
print("==================")
pprint(ministry)
print("==================")
pprint(christians)

import openpyxl 
wb = openpyxl.load_workbook('template.xlsx')

ws1 = wb.active

songCounter = 0


from tkinter import *
from tkinter import ttk

# if you are still working under a Python 2 version, 
# comment out the previous line and uncomment the following line
# import Tkinter as tk

root = Tk()
root.title('Hello world')
root.geometry('1000x500')

# Create a main frame
main_frame = Frame(root)
main_frame.pack(fill=BOTH, expand=1)

# Create a canvase
my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

# Add a scrollbar to the Canvas
my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=RIGHT, fill=Y)

# configure the canvas
my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox('all')))

# create another frame inside the canvas
second_frame = Frame(my_canvas)

my_canvas.create_window((0,0), window=second_frame, anchor="nw")


talkSpeakerDict = {}
# speakerLabels = []

def printInput(event, talk, labelRow):
    # print(test, type(event.widget))
    # print(event.widget.get())
    value = event.widget.get()
    talkSpeakerDict[talk] = value
    # speakerLabels[labelRow].config(text = value)
    pprint(talkSpeakerDict)



for i in range(0, len(talks)):
    try:
        test = re.search(".{95}", talks[i]).group()
        test += "..."
    except:
        test = talks[i]

    def handler(event, test=i, labelRow=i):
            printInput(event, talks[test], labelRow)

    Label(second_frame, text=test).grid(row=i, column=0, pady=10, padx=10)
    # speaker = Label(second_frame, text="")
    # speaker.grid(row=i, column=3, pady=10, padx=10)
    # speakerLabels.append(speaker)

    entry = (Entry(second_frame))
    entry.grid(row=i, column=1, pady=10, padx=10)
    entry.bind('<KeyRelease>', handler)



def close():
    root.quit()
Button(second_frame, text='Apply', command=close).grid(row=0, column=3)
root.mainloop()











print(ws1.max_row)
for row in range(1, ws1.max_row + 1):
    print(row)
    if(ws1[f'A{row}'].value == '[Song]'):
        ws1[f'A{row}'] = songs[songCounter]
        
        if(songCounter < len(songs) - 1):
            songCounter += 1
        

mIndex = 0
tIndex = 0
cIndex = 0

current = 'treasures'
rows = ws1.max_row


for row in range(1, 1000):
    if(ws1[f'C{row}'].value == '[Talk]'):
        if(current == 'treasures'):
            ws1[f'C{row}'] = ""
            for talk in reversed(treasures[tIndex]):
                rows += 1
                ws1.insert_rows(row, 1)
                ws1[f'C{row}'] = talk
                ws1[f'A{row}'] = talkTime[talk]   

                ws1[f'A{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                if(len(ws1[f'C{row}'].value) > 70):
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=False)


                


                try:
                    if(talkSpeakerDict[talk] != None):
                        print("speaker for ", talk)
                        ws1[f'B{row}'] = talkSpeakerDict[talk]
                        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                except: 
                    continue

            if(tIndex < len(treasures) - 1):
                tIndex += 1
            current = 'ministry'
            continue

        if(current == 'ministry'):
            ws1[f'C{row}'] = ""
            for talk in reversed(ministry[mIndex]):
                rows += 1
                ws1.insert_rows(row, 1)
                ws1[f'C{row}'] = talk
                ws1[f'A{row}'] = talkTime[talk]

                ws1[f'A{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                
                if(len(ws1[f'C{row}'].value) > 100):
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=False)

                try:
                    if(talkSpeakerDict[talk] != None):
                        print("speaker for ", talk)
                        ws1[f'B{row}'] = talkSpeakerDict[talk]
                        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                except: 
                    continue

            if(mIndex < len(ministry) - 1):
                mIndex += 1
            current = 'christians'
            continue

        if(current == 'christians'):
            ws1[f'C{row}'] = ""
            for talk in reversed(christians[cIndex]):
                rows += 1
                ws1.insert_rows(row, 1)
                ws1[f'C{row}'] = talk
                ws1[f'A{row}'] = talkTime[talk]

                ws1[f'A{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                print("LENGTH OF VALUE ", len(ws1[f'C{row}'].value))
                              
                if(len(ws1[f'C{row}'].value) > 100):
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=True)
                else:
                    ws1[f'C{row}'].alignment = openpyxl.styles.Alignment(wrap_text=False)


                try:
                    if(talkSpeakerDict[talk] != None):
                        print("speaker for ", talk)
                        ws1[f'B{row}'] = talkSpeakerDict[talk]
                        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                except: 
                    continue
            
            if(cIndex < len(christians) - 1):
                cIndex += 1
            current = 'treasures'






from datetime import timedelta
finishTime = timedelta(0, (7*3600+6*60))



# talkCell[talkTime[talk]] = f'D{row}'




# range = ws1['A1' : f'A{rows}']
# for cell in range:
#     for x in cell:
#         if("min" in x.value):






# for talk in talks:
#     try:
#         print(talkCell[talk])
#     except:
#         continue

from openpyxl.styles import Border, Side

top = Side(border_style='thin', color='FFFFFF')
left = Side(border_style='thin', color='FFFFFF')
right = Side(border_style='thin', color='FFFFFF')
bottom = Side(border_style='thin', color='FFFFFF')
border = Border(top = top, bottom = bottom, left = left, right = right)

cellRange = ws1['A1' : f'D{rows}']
for cell in cellRange:
    for x in cell:
        x.border = border
    
# right = Side(border_style='thin', color='000000')
# border = Border(top = top, bottom = bottom, left = left, right = right)

# columns = ['A', 'B', 'C', 'D']

# for i in columns:
#     cellRange = ws1[f'{i}3' : f'{i}{rows}']
#     for cell in cellRange:
#         for x in cell:
#             x.border = border

bottom = Side(border_style='thin', color='808080')
border = Border(top = top, bottom = bottom, left = left, right = right)



for row in range(1, 200):
    try:
        if('min' in ws1[f'A{row}'].value):
            stripNum = re.search('\d+', ws1[f'A{row}'].value)
            timeAdd = timedelta(0, (0*3600+int(stripNum.group())*60))
            finishTime += timeAdd
            ws1[f'D{row}'] = finishTime
            
            number_format = '[hh]:mm'
            ws1[f'D{row}'].number_format = number_format
            ws1[f'D{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            for key in talkTime:
                if(talkTime[key] == ws1[f'A{row}'].value):
                    finishTime = timedelta(0, (7*3600+6*60))

            cellRange = ws1[f'A{row}' : f'D{row}']
            for cell in cellRange:
                for x in cell:
                    x.border = border

           



            # print("Does ", ws1[f'C{row}'], " require a speaker?")
            # response = input()
            # if(response != ""):
            #     ws1[f'B{row}'] = response




    except:
        continue



first = re.sub("-.*\s", " ", book.get_metadata('DC', 'title')[0][0])
second = re.sub(",\s.*-", ", ", book.get_metadata('DC', 'title')[0][0])


ws1['B1'] = first
print(rows)

scripturesIndex = 0

for row in range(2, rows):
    if(ws1[f'B{row}'].value == '[Heading]'):
        print('here')
        ws1[f'B{row}'] = second
        ws1[f'B{row}'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        ws1.merge_cells(f'B{row}:C{row}')

    
    if(ws1[f'C{row}'].value == '[Scripture]'):
        ws1[f'C{row}'] = scriptures[scripturesIndex]
        scripturesIndex += 1



wb.save(filename = "new.xlsx")

os.startfile('new.xlsx')



